from functools import lru_cache

from  cdifflib import CSequenceMatcher
import difflib
import re

difflib.SequenceMatcher = CSequenceMatcher

from diff_match_patch import diff_match_patch
from openpyxl.workbook import Workbook

Ignored_marks = re.compile(r'\s+$')

def find_text_differences_dmp(text1: str, text2: str):
    def consists_of_whitespaces(text: str):
        return all(0 <= ord(c) <= 32 or 127 <= ord(c) <= 160 for c in text)
    dmp = diff_match_patch()


    diffs = dmp.diff_main(text1, text2)
    dmp.diff_cleanupEfficiency(diffs)
    # print(diffs)

    changes = {'added_text2': [], 'removed_text1': [], 'changed_text1': [], 'changed_text2': []}
    i = 0
    char_index_1 = 0
    char_index_2 = 0
    while i < (len(diffs) - 1):
        diff = diffs[i]
        next_diff = diffs[i + 1]
        # 0 - не изменено
        # 1 - добавлено
        # -1 - удалено
        match (diff[0], next_diff[0]):
            case (1, -1) | (-1, 1):  # Изменено
                if not consists_of_whitespaces(diff[1]) and not consists_of_whitespaces(next_diff[1]):
                    changes['changed_text1'].extend(range(char_index_1, char_index_1 + len(diff[1])))
                    changes['changed_text2'].extend(range(char_index_2, char_index_2 + len(next_diff[1])))
                char_index_1 += len(diff[1])
                char_index_2 += len(next_diff[1])
                i += 1
            case (-1, 0):  # Удалено
                if not consists_of_whitespaces(diff[1]):
                    changes['removed_text1'].extend(range(char_index_1, char_index_1 + len(diff[1])))
                char_index_1 += len(diff[1])
            case (1, 0):  # Добавлено
                if not consists_of_whitespaces(diff[1]):
                    changes['added_text2'].extend(range(char_index_2, char_index_2 + len(diff[1])))
                char_index_2 += len(diff[1])
            case (0, _):
                char_index_1 += len(diff[1])
                char_index_2 += len(diff[1])
        i += 1

    if len(diffs) > 0:
        match diffs[-1][0]:
            case 1:  # Добавлено
                changes['added_text2'].extend(range(char_index_2, char_index_2 + len(diffs[-1][1])))
                char_index_2 += len(diffs[-1][1])
            case -1:  # Удалено
                changes['removed_text1'].extend(range(char_index_1, char_index_1 + len(diffs[-1][1])))
                char_index_1 += len(diffs[-1][1])

    return changes


@lru_cache(128)
def find_text_differences_difflib(text1: str, text2: str):
    s = difflib.SequenceMatcher(lambda c: 0 <= ord(c) <= 32 or 127 <= ord(c) <= 160, text1, text2, False)
    opcodes = s.get_opcodes()

    changes = {'added_text2': [], 'removed_text1': [], 'changed_text1': [], 'changed_text2': []}
    previous_change = None
    for change in opcodes:
        from1, to1, from2, to2 = change[1], change[2], change[3], change[4]
        match change[0]:
            case 'equal':
                pass
            case 'insert':
                changes['added_text2'].extend(range(from2, to2))
            case 'replace':
                if previous_change is not None:
                    from1 = previous_change[2]
                    from2 = previous_change[4]

                changes['changed_text1'].extend(range(from1, to1))
                changes['changed_text2'].extend(range(from2, to2))
            case 'delete':
                changes['removed_text1'].extend(range(from1, to1))
        previous_change = change

    return changes


def is_empty(cell):
    if cell is None:
        return True
    return str(cell).strip() == ''


def find_spreadsheet_differences(wb1: Workbook, wb2: Workbook):
    changes = {
        'added_cells': [],
        'removed_cells': [],
        'changed_cells': [],
    }

    num_sheets = max(len(wb1.sheetnames), len(wb2.sheetnames))

    for sheet_index in range(num_sheets):
        sheet1 = wb1.worksheets[sheet_index] if sheet_index < len(wb1.sheetnames) else None
        sheet2 = wb2.worksheets[sheet_index] if sheet_index < len(wb2.sheetnames) else None

        sheet_name1 = sheet1.title if sheet1 else None
        sheet_name2 = sheet2.title if sheet2 else None

        max_row = max(sheet1.max_row if sheet1 else 0, sheet2.max_row if sheet2 else 0)
        max_col = max(sheet1.max_column if sheet1 else 0, sheet2.max_column if sheet2 else 0)

        for row_num in range(1, max_row + 1):
            for col_num in range(1, max_col + 1):
                c1 = sheet1.cell(row=row_num, column=col_num).value if sheet1 else None
                c2 = sheet2.cell(row=row_num, column=col_num).value if sheet2 else None
                if is_empty(c1) and is_empty(c2):
                    continue
                elif is_empty(c1):
                    changes['added_cells'].append((sheet_index, sheet_name1, row_num, col_num))
                    # print(
                    #     f"added_cells [sheet_index={sheet_index}, sheet_name={sheet_name2}] ({row_num}, {col_num})  {c2}")
                elif is_empty(c2):
                    changes['removed_cells'].append((sheet_index, sheet_name2, row_num, col_num))
                    # print(
                    #     f"removed_cells [sheet_index={sheet_index}, sheet_name={sheet_name1}] ({row_num}, {col_num})  {c1}")
                elif str(c1).strip() != str(c2).strip():
                    changes['changed_cells'].append((sheet_index, sheet_name1, row_num, col_num))
                    # print(
                    #     f"changed_cells [sheet_index={sheet_index}, sheet_name={sheet_name1}] ({row_num}, {col_num})  {c1} != {c2}")

    return changes
