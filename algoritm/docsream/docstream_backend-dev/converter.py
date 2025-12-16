import os

import xlrd
from openpyxl import Workbook

from config import spreadsheet_upload_folder as spreadsheet_uf


def convert_to_xlsx_workbook(xls_workbook: xlrd.book.Book) -> Workbook:
    xlsx_workbook = Workbook()

    for sheet_index in range(len(xls_workbook.sheets())):
        xls_sheet = xls_workbook.sheet_by_index(sheet_index)

        if sheet_index == 0:
            xlsx_sheet = xlsx_workbook.active
            xlsx_sheet.title = xls_sheet.name
        else:
            xlsx_sheet = xlsx_workbook.create_sheet(title=xls_sheet.name)

        for row in range(xls_sheet.nrows):
            for col in range(xls_sheet.ncols):
                cell_value = xls_sheet.cell_value(row, col)
                xlsx_sheet.cell(row=row + 1, column=col + 1, value=cell_value)

    return xlsx_workbook


def convert_xls2xlsx(filename: str) -> str:
    xls_file_path = os.path.join(spreadsheet_uf, filename)
    xls_workbook = xlrd.open_workbook(xls_file_path)

    xlsx_workbook = convert_to_xlsx_workbook(xls_workbook)

    xlsx_filename = filename.replace('.xls', '.xlsx')
    xlsx_file_path = os.path.join(spreadsheet_uf, xlsx_filename)
    xlsx_workbook.save(xlsx_file_path)

    return xlsx_filename
