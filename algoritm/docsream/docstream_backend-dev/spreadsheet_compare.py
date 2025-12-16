import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook

from config import print
from config import spreadsheet_result_folder as result_folder
from config import spreadsheet_upload_folder as upload_folder
from difference_finder import find_spreadsheet_differences


def __save_result_file_as_xlsx(book: Workbook, filename):
    book.save(filename)
    print(f"Результат сравнения сохранен в {filename}")


def highlight_cells(wb_out: Workbook, points: list[tuple[int, int]], color: str):
    sheets = wb_out.worksheets
    for (sheet_index, sheet_name, row_num, col_num) in points:
        sheets[sheet_index].cell(row=row_num, column=col_num).fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid"
        )
    return wb_out


def apply_highlights(book: Workbook, changes, color_map):
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.style = 'Normal'
    for change_type, color in color_map.items():
        highlight_cells(book, changes[change_type], color)


def compare_and_save(filename1: str, filename2: str):
    # Загрузка таблиц
    wb1 = load_workbook(os.path.join(upload_folder, filename1))
    wb2 = load_workbook(os.path.join(upload_folder, filename2))

    file_path1 = os.path.join(result_folder, filename1)
    file_path2 = os.path.join(result_folder, filename2)

    # Define colors
    yellow = "ffff00"
    red = "ff0000"
    green = "00ff00"

    # Apply highlights
    color_map1 = {
        'removed_cells': red,
        'changed_cells': yellow
    }
    color_map2 = {
        'added_cells': green,
        'changed_cells': yellow
    }
    changes = find_spreadsheet_differences(wb1, wb2)
    apply_highlights(wb1, changes, color_map1)
    apply_highlights(wb2, changes, color_map2)

    # Сохранение результатов
    __save_result_file_as_xlsx(wb1, file_path1)
    __save_result_file_as_xlsx(wb2, file_path2)
